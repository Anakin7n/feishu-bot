"""
飞书全自动 Excel 处理机器人（长连接模式）
群里发链接 → 自动下载 → 自动解析 → 自动回复
使用原生 websockets + requests，启动秒级响应，无需 lark-oapi SDK。
"""

import asyncio
import concurrent.futures
import hashlib
import inspect
import json
import logging
import os
import re
import sys
import threading
import time
import traceback
import urllib.parse
from io import BytesIO
from pathlib import Path
from urllib.parse import urlparse, parse_qs

import openpyxl
import requests
import websockets
from dotenv import load_dotenv

load_dotenv()

APP_ID = os.getenv("FEISHU_APP_ID", "")
APP_SECRET = os.getenv("FEISHU_APP_SECRET", "")

_FEISHU_DOMAIN = "https://open.feishu.cn"
_WS_ENDPOINT_URI = "/callback/ws/endpoint"

_token_cache = {"token": "", "expire": 0}


def _get_token() -> str:
    now = time.time()
    if _token_cache["token"] and now < _token_cache["expire"]:
        return _token_cache["token"]
    resp = requests.post(
        f"{_FEISHU_DOMAIN}/open-apis/auth/v3/tenant_access_token/internal",
        json={"app_id": APP_ID, "app_secret": APP_SECRET},
        timeout=10,
    )
    data = resp.json()
    if data.get("code") != 0:
        raise Exception(f"获取token失败: {data}")
    _token_cache["token"] = data["tenant_access_token"]
    _token_cache["expire"] = now + data.get("expire", 7200) - 300
    return _token_cache["token"]


def _feishu_post(path, json_body=None):
    headers = {"Authorization": f"Bearer {_get_token()}"}
    resp = requests.post(f"{_FEISHU_DOMAIN}{path}", headers=headers, json=json_body, timeout=30)
    return resp.json()
DATA_SHEET = "综拓开场数据基础模板2"

# 备用 Sheet 名和自动探测
FALLBACK_SHEET_KEYWORDS = ["综拓", "开场数据", "基础模板"]

REQUIRED_COLUMNS = [
    "场次数",
    "劣势影城数",
    "排片占比",
]

OPTIONAL_COLUMNS = [
    "影片距离满足红包场次数差1场影城数",
    "影片距离满足红包场次数差2场影城数",
]

ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS

# ---- 日志 ----

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(Path(__file__).parent / "bot.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)



# ---- 文件名解析 ----

def parse_filename(filename: str) -> dict | None:
    """从文件名解析影片名、监控日期、提取时间。"""
    name = Path(filename).stem
    t = r"(\d{2})[:,_](\d{2})(?:[:,_](\d{2}))?"
    pattern = (
        r"^(.+?)"
        r"\((\d{4}-\d{2}-\d{2})[+ ]" + t
        + r"-(\d{4}-\d{2}-\d{2})[+ ]" + t + r"\)"
        r"(\d{4}-\d{2}-\d{2})[+ ](\d{2})[:,_](\d{2})"
        r"$"
    )
    m = re.match(pattern, name)
    if not m:
        return None

    mon_start_str = f"{m.group(2)} {m.group(3)}:{m.group(4)}:{m.group(5) or '00'}"
    mon_end_str = f"{m.group(6)} {m.group(7)}:{m.group(8)}:{m.group(9) or '00'}"

    return {
        "movie": m.group(1),
        "mon_start": mon_start_str,
        "mon_end": mon_end_str,
        "extract_date": m.group(10),
        "extract_hour": int(m.group(11)),
    }


# ---- Excel 数据提取 ----

def find_data_sheet(wb) -> str | None:
    """找到包含目标列头的数据 Sheet。"""
    # 优先精确匹配
    if DATA_SHEET in wb.sheetnames:
        return DATA_SHEET

    # 按关键词模糊匹配
    for name in wb.sheetnames:
        for kw in FALLBACK_SHEET_KEYWORDS:
            if kw in name:
                return name

    # 自动探测：查找包含「场次数」列头的 sheet
    for name in wb.sheetnames:
        ws = wb[name]
        for row_idx in range(1, min(ws.max_row + 1, 10)):
            for c in range(1, min(ws.max_column + 1, 30)):
                val = str(ws.cell(row=row_idx, column=c).value or "")
                if "场次数" in val:
                    return name

    return None


def find_header_row(ws) -> int | None:
    for row_idx in range(1, min(ws.max_row + 1, 15)):
        for c in range(1, ws.max_column + 1):
            if "场次数" in str(ws.cell(row=row_idx, column=c).value or ""):
                return row_idx
    return None


def find_heji_row(ws, start_row: int) -> int | None:
    substring_match = None
    for row_idx in range(start_row + 1, ws.max_row + 1):
        val = str(ws.cell(row=row_idx, column=1).value or "").strip()
        if val == "合计":
            return row_idx
        if substring_match is None and "合计" in val:
            substring_match = row_idx
    return substring_match


def extract_all_data(ws, header_row: int, main_movie: str = "") -> dict | None:
    """提取主电影和对比电影（如有）的数据。"""
    heji_row = find_heji_row(ws, header_row)
    if heji_row is None:
        logging.warning("未找到合计行")
        return None

    # 第一遍：扫描所有列，记录每列出现次数和位置
    col_occurrences: dict[str, list[int]] = {}  # col_name -> [col_idx1, col_idx2, ...]
    for col_idx in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=col_idx).value or "").strip()
        if val:
            if val not in col_occurrences:
                col_occurrences[val] = []
            col_occurrences[val].append(col_idx)

    # 检查是否是对比模式（场次数出现 >= 2 次）
    is_cmp = "场次数" in col_occurrences and len(col_occurrences["场次数"]) >= 2

    # 映射主电影列
    main_col_map = {}
    for col_name in ALL_COLUMNS:
        if col_name in col_occurrences:
            main_col_map[col_name] = col_occurrences[col_name][0]

    missing = [c for c in REQUIRED_COLUMNS if c not in main_col_map]
    if missing:
        logging.warning(f"缺少必要列: {missing}")
        return None

    main_data = {name: ws.cell(row=heji_row, column=col).value for name, col in main_col_map.items()}

    result = {"data": main_data, "cmp_movie": None, "cmp_data": None}

    if not is_cmp:
        return result

    # 提取对比电影名：从 header 上方行中查找
    first_cmp_col = col_occurrences["场次数"][1]
    cmp_movie = None
    for r in range(header_row - 1, max(0, header_row - 4), -1):
        val = str(ws.cell(row=r, column=first_cmp_col).value or "").strip()
        if val and val not in ("", "None") and val != main_movie and "场次" not in val:
            cmp_movie = val
            break

    # 如果没在第一个对比列上方找到，扫描整行
    if not cmp_movie:
        for r in range(header_row - 1, max(0, header_row - 4), -1):
            for c in range(first_cmp_col, ws.max_column + 1):
                val = str(ws.cell(row=r, column=c).value or "").strip()
                if val and len(val) > 1 and val != main_movie and "场次" not in val and "占比" not in val and "差值" not in val:
                    cmp_movie = val
                    break
            if cmp_movie:
                break

    if not cmp_movie:
        cmp_movie = "对比影片"

    logging.info(f"对比模式: {main_movie} vs {cmp_movie}")

    # 映射对比电影列（取第 2 次出现的通用列名 + 对比专用列名）
    cmp_col_map = {}

    # 对比列中也会出现的通用列名
    cmp_shared = ["场次数", "排片占比", "劣势影城数"]
    for name in cmp_shared:
        if name in col_occurrences and len(col_occurrences[name]) >= 2:
            cmp_col_map[name] = col_occurrences[name][1]

    # 对比专用列名
    cmp_specific = [
        "场次数新增",
        "主影片与当前影片场次差值",
        "主影片与当前影片排片占比差值",
        "拍片占比",
    ]
    for name in cmp_specific:
        if name in col_occurrences:
            cmp_col_map[name] = col_occurrences[name][0]

    cmp_data = {name: ws.cell(row=heji_row, column=col).value for name, col in cmp_col_map.items()}

    # 统一 拍片占比 → 排片占比
    if "拍片占比" in cmp_data and "排片占比" not in cmp_data:
        cmp_data["排片占比"] = cmp_data.pop("拍片占比")

    result["cmp_movie"] = cmp_movie
    result["cmp_data"] = cmp_data
    return result


# ---- 文案生成 ----

def calc_deadline(extract_date: str, extract_hour: int) -> str:
    deadline_hour = 10 if extract_hour < 12 else 16
    parts = extract_date.split("-")
    return f"{int(parts[1])}月{int(parts[2])}日{deadline_hour}点"


def format_mon_date(date_str: str) -> str:
    parts = date_str.split(" ")[0].split("-")
    return f"{int(parts[1])}月{int(parts[2])}日"


def format_mon_day(date_str: str) -> str:
    return f"{int(date_str.split(' ')[0].split('-')[2])}日"


def build_message(entries: list[dict]) -> tuple[str, str]:
    movie = entries[0]["movie"]
    last = entries[-1]
    deadline = calc_deadline(last["extract_date"], last["extract_hour"])
    date_range = f"{format_mon_date(entries[0]['mon_start'])}-{format_mon_date(entries[-1]['mon_start'])}"

    # 判断是否对比模式
    has_cmp = any(e.get("cmp_movie") for e in entries)
    cmp_movie = entries[0].get("cmp_movie", "") if has_cmp else ""

    # 根据是否有对比影城拼接不同的第1条文案
    if has_cmp:
        first_line = f"1）优先推进已开预售未开《{movie}》的影城，攻克劣势影城，加速影城开场进度"
    else:
        first_line = f"1）优先推进已开预售未开《{movie}》的影城，攻克劣势影城"

    lines = [
        "辛苦同步",
        first_line,
        "2）目标进度低于均值的小伙伴们继续加油",
        "",
        f"截止至{deadline}，【{movie}】{date_range}开场数据如上",
    ]

    def _format_pp(val):
        if isinstance(val, (int, float)) and val < 1:
            return round(val * 100, 2)
        return val

    def _hb_text(d):
        parts = []
        if "影片距离满足红包场次数差1场影城数" in d:
            parts.append(f"排片红包差值1场影院:{int(d['影片距离满足红包场次数差1场影城数'])}家")
        if "影片距离满足红包场次数差2场影城数" in d:
            parts.append(f"排片红包差值2场影院:{int(d['影片距离满足红包场次数差2场影城数'])}家")
        return "，".join(parts)

    n = 1
    for entry in entries:
        day_label = format_mon_day(entry["mon_start"])
        d = entry["data"]
        cc = int(d["场次数"])
        ls = int(d["劣势影城数"])
        pp = _format_pp(d["排片占比"])

        # 主电影行
        line = f"{n}）{day_label}《{movie}》已开{cc}场，未排《{movie}》的有{ls}家，排片占比{pp}%"
        hb = _hb_text(d)
        if hb:
            line += "，" + hb
        lines.append(line)
        n += 1

        # 对比电影行
        if has_cmp and entry.get("cmp_data"):
            cd = entry["cmp_data"]
            ccc = int(cd["场次数"])
            cpp = _format_pp(cd["排片占比"])
            diff = cd.get("主影片与当前影片排片占比差值")
            if diff is None and isinstance(d["排片占比"], (int, float)) and isinstance(cd["排片占比"], (int, float)):
                diff = round((_format_pp(d["排片占比"]) - _format_pp(cd["排片占比"])), 2)
            elif isinstance(diff, (int, float)) and diff < 1 and diff != 0:
                diff = round(diff * 100, 2)

            cmp_line = f"{n}）{day_label}《{cmp_movie}》已开{ccc}场，排片占比{cpp}%"
            if diff is not None:
                cmp_line += f"，{movie}与{cmp_movie}大盘差值为{diff}%"
            lines.append(cmp_line)
            n += 1

        lines.append("")

    main_message = "\n".join(lines)

    # 第二条消息：总结跟进语
    summary_message = f"以上为截止{deadline}，《{movie}》{date_range}开场情况，辛苦大家参考跟进。"

    return main_message, summary_message


# ---- 文件下载 ----

def download_file(url: str) -> tuple[str | None, bytes | None]:
    """下载 Excel，返回 (文件名, 内容)。"""
    try:
        resp = requests.get(url, timeout=60, headers={"Accept-Encoding": "identity"})
        resp.raise_for_status()

        fname = None
        cd = resp.headers.get("Content-Disposition", "")
        if "filename*=" in cd:
            parts = cd.split("filename*=")
            if len(parts) > 1:
                encoded = parts[1].split(";")[0].strip()
                if "''" in encoded:
                    _, fname_encoded = encoded.split("''", 1)
                    fname = urllib.parse.unquote(fname_encoded)
        if not fname and "filename=" in cd:
            raw = cd.split("filename=")[1].split(";")[0].strip().strip('"')
            try:
                fname = raw.encode("latin-1").decode("utf-8")
            except (UnicodeDecodeError, UnicodeEncodeError):
                fname = raw
        if not fname:
            fname = url.split("/")[-1].split("?")[0]

        return fname, resp.content
    except Exception as e:
        logging.error(f"下载失败: {e}")
        return None, None


# ---- 飞书文件下载 ----

def download_feishu_file(message_id: str, file_key: str) -> tuple[str | None, bytes | None, str | None]:
    """下载飞书消息中的文件附件，返回 (文件名, 内容, 错误信息)。"""
    def _parse_feishu_error(resp) -> str:
        """从飞书错误响应中提取 msg，提取不到返回空字符串。"""
        try:
            body = resp.json()
            return f"（{body.get('msg', '')}）"
        except Exception:
            return ""

    try:
        resp = requests.get(
            f"{_FEISHU_DOMAIN}/open-apis/im/v1/messages/{message_id}/resources/{file_key}",
            params={"type": "file"},
            headers={"Authorization": f"Bearer {_get_token()}"},
            timeout=60,
        )

        if resp.status_code == 200:
            fname = None
            cd = resp.headers.get("Content-Disposition", "")

            # RFC 5987: filename*=UTF-8''percent-encoded
            if "filename*=" in cd:
                parts = cd.split("filename*=")
                if len(parts) > 1:
                    encoded = parts[1].split(";")[0].strip()
                    if "''" in encoded:
                        _, fname_encoded = encoded.split("''", 1)
                        fname = urllib.parse.unquote(fname_encoded)

            # Fallback: filename="..."
            if not fname and "filename=" in cd:
                raw = cd.split("filename=")[1].split(";")[0].strip().strip('"')
                # 飞书可能把 UTF-8 字节直接塞进 filename= 字段，
                # requests 按 HTTP 标准用 Latin-1 解码了头字段，导致中文变乱码。
                # 尝试 Latin-1 → bytes → UTF-8 还原。
                try:
                    fname = raw.encode("latin-1").decode("utf-8")
                except (UnicodeDecodeError, UnicodeEncodeError):
                    fname = raw

            if not fname:
                fname = file_key + ".xlsx"
            return fname, resp.content, None

        # 非 200：解析飞书错误信息
        feishu_err = _parse_feishu_error(resp)
        if resp.status_code == 404:
            return None, None, f"文件已过期或不存在{feishu_err}，请重新发送"
        if resp.status_code == 403:
            return None, None, f"权限不足{feishu_err}，请确认应用已开通 im:resource 权限"
        if resp.status_code == 400:
            return None, None, f"文件无法下载{feishu_err}，文件可能已失效或不支持下载（如转发消息中的文件）"
        return None, None, f"下载失败（HTTP {resp.status_code}{feishu_err}），请重试"
    except requests.Timeout:
        return None, None, "下载超时，请重试"
    except requests.RequestException as e:
        logging.error(f"下载飞书文件失败: {e}")
        return None, None, "网络异常，下载失败，请重试"
    except Exception as e:
        logging.error(f"下载飞书文件失败: {e}")
        return None, None, "下载失败，请重试"


# ---- 消息发送 ----

def send_message(chat_id: str, text: str):
    content = json.dumps({"text": text}, ensure_ascii=False)
    resp = _feishu_post(
        f"/open-apis/im/v1/messages?receive_id_type=chat_id",
        {"receive_id": chat_id, "msg_type": "text", "content": content},
    )
    if resp.get("code") == 0:
        logging.info("已发送到群聊")
    else:
        logging.error(f"发送失败 code={resp.get('code')}, msg={resp.get('msg')}")


# 消息去重（内存缓存 + 文件持久化，重启不丢失）
_SEEN_FILE = Path(__file__).parent / ".seen_msg_ids"
_SEEN_MAX = 500  # 触发裁剪的阈值
_SEEN_KEEP = 300  # 裁剪时保留最近条数

_seen_cache: set[str] | None = None
_seen_lock = threading.RLock()


def _init_seen():
    global _seen_cache
    _seen_cache = set()
    if _SEEN_FILE.exists():
        with open(_SEEN_FILE, "r") as f:
            for line in f:
                stripped = line.strip()
                if stripped:
                    _seen_cache.add(stripped)


def _trim_seen():
    with _seen_lock:
        if not _SEEN_FILE.exists():
            return
        with open(_SEEN_FILE, "r") as f:
            all_lines = [line.strip() for line in f if line.strip()]
        if len(all_lines) <= _SEEN_MAX:
            return
        kept = all_lines[-_SEEN_KEEP:]
        with open(_SEEN_FILE, "w") as f:
            for line in kept:
                f.write(line + "\n")
        global _seen_cache
        _seen_cache = set(kept)


def _is_duplicate(msg_id: str) -> bool:
    global _seen_cache
    with _seen_lock:
        if _seen_cache is None:
            _init_seen()
        if msg_id in _seen_cache:
            return True
        _seen_cache.add(msg_id)
        with open(_SEEN_FILE, "a") as f:
            f.write(msg_id + "\n")
            f.flush()
            os.fsync(f.fileno())
        if len(_seen_cache) > _SEEN_MAX:
            _trim_seen()
        return False


_content_seen: dict[str, float] = {}
_CONTENT_TTL = 600  # 内容哈希 10 分钟后过期


def _is_content_duplicate(content_str: str) -> bool:
    with _seen_lock:
        now = time.time()
        h = hashlib.sha256(content_str.encode()).hexdigest()
        if h in _content_seen and now - _content_seen[h] < _CONTENT_TTL:
            return True
        _content_seen[h] = now
        return False


# ---- 文件消息累积 ----

_pending_file: dict[str, dict] = {}  # chat_id -> entry (等待配对的第一个文件)
_pending_timers: dict[str, threading.Timer] = {}
_pending_lock = threading.Lock()
_FILE_WAIT_SECONDS = 5  # 等第二个文件的超时时间


def _on_file_timeout(chat_id: str):
    """5 秒内没收到第二个文件，提醒用户（文件继续暂存）。"""
    logging.info(f"[文件提醒] chat_id={chat_id} 未收到第二个文件，继续等待")
    send_message(chat_id, "仍需要一个 Excel 文件，请继续发送第二个文件")
    # 重新计时
    with _pending_lock:
        if chat_id in _pending_file:
            timer = threading.Timer(_FILE_WAIT_SECONDS, _on_file_timeout, args=[chat_id])
            timer.daemon = True
            _pending_timers[chat_id] = timer
            timer.start()


def _process_two_files(chat_id: str, entry1: dict, entry2: dict):
    """两个文件到齐，合并生成文案。"""
    entries = [entry1, entry2]
    entries.sort(key=lambda e: e["mon_start"])
    logging.info("两个文件到齐，生成文案")
    main_msg, summary_msg = build_message(entries)
    send_message(chat_id, main_msg)
    time.sleep(0.5)
    send_message(chat_id, summary_msg)


# ---- 事件处理 ----

def _parse_single_file(fname: str, content: bytes) -> dict | None:
    """解析单个 Excel 文件，返回 entry dict 或 None。"""
    safe_name = re.sub(r'[<>:"/\\|?*]', '_', fname)
    logging.info(f"文件: {safe_name}")

    info = parse_filename(safe_name)
    if info is None:
        logging.warning(f"文件名格式不匹配: {safe_name}")
        return None

    logging.info(f"影片: {info['movie']}  监控: {info['mon_start']} ~ {info['mon_end']}")

    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        logging.warning(f"Excel 文件打开失败: {e}")
        return None

    sheet_name = find_data_sheet(wb)
    if sheet_name is None:
        logging.warning(f"未找到数据 Sheet，可用: {wb.sheetnames}")
        wb.close()
        return None

    ws = wb[sheet_name]
    logging.info(f"Sheet: {sheet_name}  (max_row={ws.max_row}, max_col={ws.max_column})")
    header_row = find_header_row(ws)
    if header_row is None:
        logging.warning("未找到表头")
        wb.close()
        return None

    logging.info(f"表头行: {header_row}")

    all_data = extract_all_data(ws, header_row, info["movie"])
    wb.close()

    if all_data is None:
        logging.warning("数据提取失败")
        return None

    info["data"] = all_data["data"]
    if all_data["cmp_movie"]:
        info["cmp_movie"] = all_data["cmp_movie"]
        info["cmp_data"] = all_data["cmp_data"]

    logging.info(f"数据: {info['data']}")
    if info.get("cmp_movie"):
        logging.info(f"对比: {info['cmp_movie']} -> {info['cmp_data']}")
    return info


def process_urls(urls: list[str]) -> tuple[str, str] | None:
    """下载链接中的 Excel 并解析，至少需要 2 个有效文件。"""
    logging.info("开始处理链接...")
    entries = []

    for url in urls:
        logging.info(f"下载: {url}")
        fname, content = download_file(url)
        if not content:
            logging.warning("下载失败")
            continue

        entry = _parse_single_file(fname, content)
        if entry:
            entries.append(entry)

    if len(entries) < 2:
        logging.warning(f"至少需要 2 个有效文件，当前只有 {len(entries)} 个")
        return None

    entries.sort(key=lambda e: e["mon_start"])
    logging.info("处理完成，生成文案")
    return build_message(entries)


def on_message_receive(event_data: dict) -> None:
    event = event_data.get("event", {})
    msg = event.get("message", {})
    msg_id = msg.get("message_id", "")

    if _is_duplicate(msg_id):
        return

    create_time_ms = msg.get("create_time", "")
    if create_time_ms:
        try:
            age = time.time() - int(create_time_ms) / 1000
            if age > 180:
                logging.info(f"[跳过] 消息过旧 ({age:.0f}s 前)")
                return
        except (ValueError, OSError):
            pass

    chat_id = msg.get("chat_id", "")
    msg_type = msg.get("message_type", "")

    sender = event.get("sender", {})
    if sender.get("sender_type") == "app":
        return

    logging.info(f"[事件] chat_id={chat_id}  msg_type={msg_type}")

    content_str = msg.get("content", "{}")

    # ---- 文件消息处理 ----
    if msg_type == "file":
        try:
            content_json = json.loads(content_str)
            file_key = content_json.get("file_key", "")
        except (json.JSONDecodeError, AttributeError):
            file_key = ""

        if not file_key:
            logging.info("[跳过] 文件消息缺少 file_key")
            return

        if _is_content_duplicate(content_str):
            logging.info(f"[跳过] 内容重复 (file)")
            return

        logging.info(f"[文件消息] file_key={file_key}")
        fname, content, err = download_feishu_file(msg_id, file_key)
        if err:
            logging.warning(f"文件下载失败: {err}")
            send_message(chat_id, err)
            return

        # 预检：确保能解析
        entry = _parse_single_file(fname, content)
        if entry is None:
            send_message(chat_id, "文件处理失败：请确认文件名格式正确，且包含数据 Sheet")
            return

        # 检查是否有待配对的第一个文件
        first_entry = None
        timer_to_cancel = None
        with _pending_lock:
            if chat_id in _pending_file:
                first_entry = _pending_file.pop(chat_id)
                timer_to_cancel = _pending_timers.pop(chat_id, None)

        if timer_to_cancel:
            timer_to_cancel.cancel()

        if first_entry is not None:
            # 第二个文件到齐，立即处理
            _process_two_files(chat_id, first_entry, entry)
            return

        # 第一个文件：暂存，启动 5 秒超时
        with _pending_lock:
            _pending_file[chat_id] = entry
            timer = threading.Timer(_FILE_WAIT_SECONDS, _on_file_timeout, args=[chat_id])
            timer.daemon = True
            _pending_timers[chat_id] = timer
            timer.start()

        logging.info(f"[文件累积] chat_id={chat_id} 收到第 1 个文件，等待第 2 个（{_FILE_WAIT_SECONDS}s 超时）")
        return

    if msg_type != "text":
        logging.info(f"[跳过] 非文本/文件消息: {msg_type}")
        return

    logging.info(f"[内容] {content_str}")

    try:
        content_json = json.loads(content_str)
        text = content_json.get("text", "")
    except (json.JSONDecodeError, AttributeError):
        text = content_str

    if not text:
        logging.info("[跳过] 空文本")
        return

    text = re.sub(r'@\S+\s*', '', text).strip()
    logging.info(f"[文本] {text}")

    # 检测 "test" 前缀：testhttps://... 或 test https://... 形式跳过内容去重
    skip_content_dedup = bool(re.search(r'test\s*https?://', text, re.IGNORECASE))
    if skip_content_dedup:
        text = re.sub(r'test\s*(https?://)', r'\1', text, flags=re.IGNORECASE)
        logging.info("[test模式] 跳过内容去重")

    if not skip_content_dedup and _is_content_duplicate(content_str):
        logging.info(f"[跳过] 内容重复")
        return

    text = text.replace("；", " ").replace("，", " ").replace(",", " ").replace("\n", " ")
    urls = re.findall(r"https?://[^\s]+", text)
    if not urls:
        logging.info("[跳过] 未发现链接")
        return

    logging.info(f"收到 {len(urls)} 个链接: {urls}")
    result = process_urls(urls)
    if result:
        logging.info("文案已生成")
        main_msg, summary_msg = result
        send_message(chat_id, main_msg)
        time.sleep(0.5)
        send_message(chat_id, summary_msg)
    else:
        logging.warning("处理失败，发送错误提示")
        send_message(chat_id, "处理失败：请确认两个链接的 Excel 文件都包含正确的数据 Sheet")


# ---- 简化 protobuf frame 编解码 ----

def _pb_varint_size(n: int) -> int:
    size = 1
    while n > 127:
        size += 1
        n >>= 7
    return size


def _pb_write_varint(buf: bytearray, n: int):
    while n > 127:
        buf.append((n & 0x7F) | 0x80)
        n >>= 7
    buf.append(n & 0x7F)


def _pb_write_tag(buf: bytearray, field: int, wire: int):
    _pb_write_varint(buf, (field << 3) | wire)


def _pb_write_bytes(buf: bytearray, data: bytes):
    _pb_write_varint(buf, len(data))
    buf.extend(data)


def _pb_write_string(buf: bytearray, field: int, value: str):
    _pb_write_tag(buf, field, 2)
    encoded = value.encode("utf-8")
    _pb_write_bytes(buf, encoded)


def _pb_write_uint64(buf: bytearray, field: int, value: int):
    _pb_write_tag(buf, field, 0)
    _pb_write_varint(buf, value)


def _pb_write_int32(buf: bytearray, field: int, value: int):
    _pb_write_tag(buf, field, 0)
    _pb_write_varint(buf, value)


def encode_ping_frame(service_id: int) -> bytes:
    buf = bytearray()
    header_buf = bytearray()
    _pb_write_string(header_buf, 1, "type")
    _pb_write_string(header_buf, 2, "ping")
    _pb_write_int32(buf, 4, 0)
    _pb_write_uint64(buf, 2, 0)
    _pb_write_uint64(buf, 1, 0)
    _pb_write_tag(buf, 3, 0)
    _pb_write_varint(buf, service_id)
    _pb_write_tag(buf, 5, 2)
    _pb_write_varint(buf, len(header_buf))
    buf.extend(header_buf)
    return bytes(buf)


def decode_frame(data: bytes) -> dict:
    result = {}
    pos = 0
    while pos < len(data):
        if pos >= len(data):
            break
        tag = data[pos]
        pos += 1
        field = tag >> 3
        wire = tag & 0x07
        if wire == 0:
            value = 0
            shift = 0
            while pos < len(data):
                b = data[pos]
                pos += 1
                value |= (b & 0x7F) << shift
                shift += 7
                if not (b & 0x80):
                    break
            result[field] = value
        elif wire == 2:
            length = 0
            shift = 0
            while pos < len(data):
                b = data[pos]
                pos += 1
                length |= (b & 0x7F) << shift
                shift += 7
                if not (b & 0x80):
                    break
            value = data[pos:pos + length]
            pos += length
            if field == 5:
                headers = []
                hpos = 0
                while hpos < len(value):
                    htag = value[hpos]
                    hpos += 1
                    hfield = htag >> 3
                    hwire = htag & 0x07
                    if hwire == 2:
                        hlen = 0
                        hshift = 0
                        while hpos < len(value):
                            hb = value[hpos]
                            hpos += 1
                            hlen |= (hb & 0x7F) << hshift
                            hshift += 7
                            if not (hb & 0x80):
                                break
                        headers.append((hfield, value[hpos:hpos + hlen].decode("utf-8")))
                        hpos += hlen
                result[field] = headers
            else:
                result[field] = value
    return result


def frame_type(frame: dict) -> int:
    return frame.get(4, -1)


def frame_data_payload(frame: dict) -> bytes:
    return frame.get(8, b"")


# ---- 精简 WebSocket 客户端 ----

class FeishuWsClient:
    def __init__(self, app_id: str, app_secret: str):
        self.app_id = app_id
        self.app_secret = app_secret
        self.service_id = ""
        self._reconnect_interval = 120
        self._ping_interval = 120
        self._ws = None
        self._ws_url = ""
        self._ping_task = None
        self._executor = concurrent.futures.ThreadPoolExecutor(max_workers=2)
        self._reconnecting = False

    def _get_ws_url(self):
        resp = requests.post(
            f"{_FEISHU_DOMAIN}{_WS_ENDPOINT_URI}",
            headers={"locale": "zh"},
            json={"AppID": self.app_id, "AppSecret": self.app_secret},
            timeout=30,
        )
        data = resp.json()
        if data.get("code") != 0:
            raise Exception(f"获取WS地址失败: {data}")
        dd = data.get("data", {})
        if dd.get("ClientConfig"):
            cc = dd["ClientConfig"]
            self._reconnect_interval = cc.get("ReconnectInterval", 120)
            self._ping_interval = cc.get("PingInterval", 120)
        return dd["URL"]

    async def _ping_loop(self):
        while True:
            try:
                if self._ws is not None:
                    sid = int(self.service_id) if self.service_id else 0
                    ping = encode_ping_frame(sid)
                    await self._ws.send(ping)
            except Exception as e:
                logging.warning(f"[ping失败] {e}")
            await asyncio.sleep(self._ping_interval)

    def _dispatch_sync(self, event_data: dict):
        try:
            on_message_receive(event_data)
        except Exception as e:
            logging.error(f"[分发异常] {e}")
            traceback.print_exc()

    async def _process_event(self, event_data: dict):
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(self._executor, self._dispatch_sync, event_data)

    async def _read_loop(self):
        while True:
            try:
                raw = await self._ws.recv()
                if isinstance(raw, str):
                    continue
                frame = decode_frame(raw)
                ft = frame_type(frame)
                if ft == 0:
                    continue
                elif ft == 1:
                    payload = frame_data_payload(frame)
                    if not payload:
                        continue
                    event_data = json.loads(payload.decode("utf-8"))
                    asyncio.create_task(self._process_event(event_data))
            except websockets.exceptions.ConnectionClosed:
                logging.warning("[连接断开]")
                break
            except Exception as e:
                logging.error(f"[读取异常] {e}")
                traceback.print_exc()

    async def _try_connect(self):
        url = self._get_ws_url()
        u = urlparse(url)
        q = parse_qs(u.query)
        self.service_id = q.get("service_id", [""])[0]
        logging.info(f"[WS地址] {url[:80]}...")
        logging.info(f"[服务ID] {self.service_id}")

        params = inspect.signature(websockets.connect).parameters
        kwargs = {"proxy": None} if "proxy" in params else {}
        self._ws = await websockets.connect(url, **kwargs)
        self._ws_url = url
        logging.info("[WS已连接]")
        self._reconnecting = False
        self._ping_task = asyncio.create_task(self._ping_loop())
        await self._read_loop()

    async def connect(self):
        while True:
            try:
                await self._try_connect()
            except Exception as e:
                logging.error(f"[连接失败] {e}")
            if self._ws is not None:
                await self._ws.close()
                self._ws = None
            if self._ping_task is not None:
                self._ping_task.cancel()
                self._ping_task = None
            self._reconnecting = True
            logging.info(f"[重连] {self._reconnect_interval}s 后重试...")
            await asyncio.sleep(self._reconnect_interval)

    def start(self):
        asyncio.run(self.connect())


# ---- 主入口 ----

def main():
    # CLI 模式：直接传文件路径，不启动 WebSocket
    if len(sys.argv) > 1:
        file_list = []
        for arg in sys.argv[1:]:
            p = Path(arg)
            if not p.exists():
                print(f"文件不存在: {arg}")
                sys.exit(1)
            with open(p, "rb") as f:
                content = f.read()
            file_list.append((p.name, content))

        entries = []
        for fname, content in file_list:
            entry = _parse_single_file(fname, content)
            if entry:
                entries.append(entry)

        if not entries:
            print("没有成功解析的文件")
            sys.exit(1)

        entries.sort(key=lambda e: e["mon_start"])
        main_msg, summary_msg = build_message(entries)
        print(main_msg)
        print("---")
        print(summary_msg)
        return

    if not APP_ID or not APP_SECRET:
        logging.error("请先在 .env 中配置 FEISHU_APP_ID 和 FEISHU_APP_SECRET")
        sys.exit(1)

    logging.info(f"启动机器人 (App ID: {APP_ID})...")
    logging.info("监听群聊消息中，可直接发送链接到群里测试")

    client = FeishuWsClient(APP_ID, APP_SECRET)
    client.start()


if __name__ == "__main__":
    main()
